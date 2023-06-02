import tkinter as tk
from tkinter import filedialog
import fitz
from openpyxl import Workbook
import re

def dosya_ac():
    dosya_yolu = filedialog.askopenfilename(initialdir="/", title="Dosya Seç", filetypes=(("PDF Dosyaları", "*.pdf"), ("Tüm Dosyalar", "*.*")))
    dosya_yolu_etiketi.config(text="Seçilen Dosya: " + dosya_yolu)
    verileri_listele(dosya_yolu)

def verileri_listele(dosya_yolu):
    wb = Workbook()
    ws = wb.active

    pdf = fitz.open(dosya_yolu)
    metin = ""

    for sayfa in pdf:
        metin += sayfa.get_text()

    satirlar = metin.strip().split("\n")
    basliklar = satirlar[0].split("\t")

    basliklar = [baslik for baslik in basliklar if baslik not in ["Kelime Raporu", "ANAHTAR KELİMELER", "SIRA"]]

    # Başlıkları Excel sayfasına ekle
    for col, baslik in enumerate(basliklar, start=1):
        ws.cell(row=1, column=1, value=baslik)

    # Verileri Excel sayfasına ekle
    for row, satir in enumerate(satirlar[1:], start=2):
        veri = satir.split("\t")
        for col, sütun in enumerate(veri, start=1):
            if not any(char in sütun for char in ["(", ")", ".","/", "SIRA"]):
                # Değişiklik: Sayı içeren sütunları bul
                if re.match(r'^\d+$', sütun):
                    sayfa, sira = sayfa_sira_bul(int(sütun))
                    if sütun.endswith('"'):
                        ws.cell(row=row-2, column=2, value='{}.Sayfa,{}.Sıra"'.format(sayfa, sira))
                    else:
                        ws.cell(row=row-2, column=2, value='{}.Sayfa,{}.Sıra'.format(sayfa, sira))
                else:
                    ws.cell(row=row, column=col, value=sütun)        
    dosya_adi = dosya_yolu.split("/")[-1].split(".")[0] +"_siralama.xlsx"
    dosya_yolu = filedialog.asksaveasfilename(initialdir="/", title="Dosya Kaydet", initialfile=dosya_adi, filetypes=(("Excel Dosyaları", "*.xlsx"), ("Tüm Dosyalar", "*.*")))
    if dosya_yolu:
        wb.save(dosya_yolu)
        tk.messagebox.showinfo("Başarılı", "Excel dosyası kaydedildi.")

def sayfa_sira_bul(sayi):
    sayfa = (sayi - 1) // 10 + 1
    sira = (sayi - 1) % 10 + 1
    return sayfa, sira

# Ana pencereyi oluştur
pencere = tk.Tk()
pencere.title("PDF Dosyası Seçme Programı")
pencere.geometry("600x400")

# Dosya açma düğmesini oluştur
dosya_ac_dugme = tk.Button(pencere, text="PDF Dosyası Aç", command=dosya_ac)
dosya_ac_dugme.pack(pady=10)
dosya_yolu_etiketi = tk.Label(pencere, text="Seçilen Dosya: ")
dosya_yolu_etiketi.pack()
pencere.mainloop()
